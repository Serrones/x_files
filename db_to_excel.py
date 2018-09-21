import os
import os.path
import mysql.connector
from openpyxl import Workbook, load_workbook

mydb = mysql.connector.connect(
    host="localhost",
    user="user_excel",
    passwd="password_excel",
    database="database-teste"
    )

mycursor = mydb.cursor()

# Select all tables'
mycursor.execute("SHOW TABLES")

tables = []
for item in mycursor:
    tables.append(item[0])

for table in tables:
    # Select table
    sql = "SELECT * FROM " + table
    mycursor.execute(sql)

    myresult = mycursor.fetchall()

    # Verify if file already exists
    current_directory = os.getcwd()
    excel_file = current_directory + '/tabelas_database.xlsx'
    i = 0
    if os.path.isfile(excel_file):
        i += 1
        wb = load_workbook(excel_file)
        sheet = wb.create_sheet(index=i)
        if len(table) < 30:
            sheet.title = table
        else:
            tabela = b = 'Tabela ' + table[0:20] + '...'
            sheet.title = tabela
        sheet.append(mycursor.column_names)
        for row in myresult:
            sheet.append(row)
        wb.save(excel_file)

    else:
        # Create Excel (.xlsx) file
        wb = Workbook()
        sheet = wb.create_sheet(i)
        if len(table) < 30:
            sheet.title = table
        else:
            tabela = b = 'Tabela ' + table[0:20] + '...'
            sheet.title = tabela
        sheet.append(mycursor.column_names)
        for row in myresult:
            sheet.append(row)

        workbook_name = "tabelas_database"
        wb.save(workbook_name + ".xlsx")
